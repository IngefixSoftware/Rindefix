from fastapi import FastAPI, Depends, UploadFile, File, Form, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import text
import os
from pathlib import Path
from uuid import uuid4
from datetime import datetime
import json
import smtplib
from email.message import EmailMessage
from io import BytesIO
from openpyxl import load_workbook

from database import get_db, engine
import models
import schemas
from schemas import ReportEmailRequest
import crud

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", BASE_DIR / "uploads"))
UPLOAD_DIR.mkdir(exist_ok=True)
EXCEL_TEMPLATE_PATH = BASE_DIR / "plantillas" / "FORMATO REND.xlsx"

models.Base.metadata.create_all(bind=engine)

# Ensure extra columns for "fondo por rendir" exist in DB (PostgreSQL on Render, etc.).
try:
    with engine.begin() as conn:
        conn.execute(
            text(
                "ALTER TABLE users "
                "ADD COLUMN IF NOT EXISTS fondo_por_rendir_assigned NUMERIC(14,2) NOT NULL DEFAULT 0"
            )
        )
        conn.execute(
            text(
                "ALTER TABLE users "
                "ADD COLUMN IF NOT EXISTS fondo_por_rendir_available NUMERIC(14,2) NOT NULL DEFAULT 0"
            )
        )
        conn.execute(
            text(
                "ALTER TABLE users "
                "ADD COLUMN IF NOT EXISTS fondo_por_rendir_overdraft_limit NUMERIC(14,2) NOT NULL DEFAULT 0"
            )
        )
except Exception:
    # If the database backend doesn't support this syntax or table doesn't exist yet,
    # fail silently; on compatible backends (Render PostgreSQL) it will succeed.
    pass

app = FastAPI(
    title="RindeFix API",
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
)

# CORS para que el frontend pueda llamar al backend
origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "https://rindefix.onrender.com",
    "https://rindefixbackend.onrender.com",
]

allow_all_origins = os.getenv("CORS_ALLOW_ALL", "true").lower() == "true"
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"] if allow_all_origins else origins,
    allow_credentials=not allow_all_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

DOCUMENT_TYPES = {"FACTURA", "BOLETA", "BOLETA_COMBUSTIBLE", "COMPROBANTE_RECIBO_TBK"}
RENDITION_TYPES = {"CAJA_CHICA", "COMBUSTIBLE", "FONDO_POR_RENDIR"}
ALLOWED_ROLES = {"RENDIDOR", "APROBADOR", "ADMIN"}
BRANCHES = {
    "Rancagua",
    "Concepción",
    "Coquimbo",
    "Viña del Mar",
    "Temuco",
    "Casa Matriz",
}

SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_RECIPIENT = os.getenv("EMAIL_RECIPIENT")

DEFAULT_EMAIL_SUBJECT = (
    "Informe de rendiciones pendientes - {{rendidor_nombre}} "
    "(Total: {{monto_total_pendiente}})"
)
DEFAULT_EMAIL_BODY = (
    "Estimado,\n\n"
    "Adjunto encontrarás el informe de rendiciones pendientes del usuario {{rendidor_nombre}}.\n"
    "El monto total de las rendiciones pendientes de este informe es {{monto_total_pendiente}}.\n\n"
    "Saludos."
)


def parse_date_param(value: str, field: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Fecha inválida para {field} (YYYY-MM-DD)")


def validate_role(role: str):
    if role not in ALLOWED_ROLES:
        raise HTTPException(status_code=400, detail="Rol inválido")


def validate_branch(branch: str):
    if branch not in BRANCHES:
        raise HTTPException(status_code=400, detail="Sucursal inválida")


def format_chilean_peso(value: float) -> str:
    """
    Formatea un número como peso chileno, por ejemplo:
    1234567.8 -> "$1.234.568"
    (redondea al peso más cercano, sin decimales).
    """
    try:
      rounded = int(round(float(value)))
    except Exception:
      return str(value)
    parts = f"{rounded:,}".replace(",", ".")
    return f"${parts}"


def _safe_set(ws, coord: str, value):
    """
    Escribe en una celda evitando crashear si es parte de un merge
    (MergedCell es de solo lectura). En ese caso, simplemente se omite.
    """
    try:
        ws[coord].value = value
    except AttributeError:
        # Celda combinada de solo lectura; la plantilla manda.
        return


def _build_pending_excel(
    expenses: list[models.Expense],
    user: models.User,
    is_fondo_por_rendir: bool = False,
) -> bytes:
    template_path = EXCEL_TEMPLATE_PATH

    if not template_path.exists():
        raise HTTPException(
            status_code=500,
            detail="Plantilla de Excel no encontrada en el servidor.",
        )

    try:
        wb = load_workbook(template_path)
    except Exception as exc:
        # Error típico cuando el archivo no es un .xlsx válido
        raise HTTPException(
            status_code=500,
            detail=(
                f"Error al abrir la plantilla de Excel '{template_path.name}'. "
                "Asegúrate de que es un archivo .xlsx válido. "
                f"Detalle técnico: {exc}"
            ),
        )
    ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active

    # Fecha de rendición (fecha de generación del informe) en C13
    _safe_set(ws, "C13", datetime.now().date())

    # Datos del responsable
    # GH12: nombre del rendidor
    # GH13: sucursal del rendidor
    # GH14: fondo asignado al rendidor
    _safe_set(ws, "G12", user.name)
    _safe_set(ws, "G13", user.branch)
    assigned_budget = (
        float(user.fondo_por_rendir_assigned or 0)
        if is_fondo_por_rendir
        else float(user.budget_assigned or 0)
    )
    _safe_set(ws, "G14", assigned_budget)

    invoices = [e for e in expenses if e.document_type == "FACTURA"]
    others = [e for e in expenses if e.document_type != "FACTURA"]

    # Sección de facturas (tabla superior)
    start_row_invoices = 21
    for idx, e in enumerate(invoices):
        row = start_row_invoices + idx
        _safe_set(ws, f"C{row}", e.expense_date)
        _safe_set(ws, f"D{row}", e.document_number or "")
        # La plantilla tiene E:F:G combinadas como un solo campo de detalle.
        # Escribimos todo el detalle solo en la columna E para respetar el merge.
        detail_parts = [e.provider or "", e.title or "", e.description or ""]
        detail = " - ".join(part for part in detail_parts if part)
        _safe_set(ws, f"E{row}", detail)
        _safe_set(ws, f"H{row}", float(e.amount or 0))

    # Sección de otros documentos (tabla inferior)
    start_row_others = 37
    for idx, e in enumerate(others):
        row = start_row_others + idx
        _safe_set(ws, f"C{row}", e.expense_date)
        _safe_set(ws, f"D{row}", e.document_type or "")
        _safe_set(ws, f"E{row}", e.document_number or "")
        other_detail_parts = [e.provider or "", e.title or "", e.description or ""]
        other_detail = " - ".join(part for part in other_detail_parts if part)
        # En la plantilla, F:G suelen estar combinadas como un solo campo de detalle.
        # Escribimos el detalle solo en F para respetar el merge.
        _safe_set(ws, f"F{row}", other_detail)
        _safe_set(ws, f"H{row}", float(e.amount or 0))

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _build_report_csv(expenses: list[models.Expense]) -> str:
    from io import StringIO
    import csv

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Proveedor", "Título", "N° Documento", "Estado", "Monto"])
    total = 0.0
    for e in expenses:
        total += float(e.amount or 0)
        writer.writerow(
            [
                e.expense_date.isoformat() if e.expense_date else "",
                e.provider or "",
                e.title or "",
                e.document_number or "",
                e.status or "",
                float(e.amount or 0),
            ]
        )
    writer.writerow(["", "", "", "", "Total", total])
    return output.getvalue()


def _send_report_email(file_content: bytes, filename: str, subject: str, body: str):
    if not (SMTP_HOST and EMAIL_SENDER and EMAIL_RECIPIENT):
        raise HTTPException(
            status_code=500,
            detail="Envío de correos no está configurado en el servidor.",
        )
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = EMAIL_SENDER
    msg["To"] = EMAIL_RECIPIENT
    msg.set_content(body)
    msg.add_attachment(
        file_content,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename if filename.endswith(".xlsx") else f"{filename}.xlsx",
    )
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            if SMTP_USER and SMTP_PASSWORD:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    except Exception:
        raise HTTPException(status_code=500, detail="No se pudo enviar el correo.")


def _render_email_template(
    db: Session, rendidor_name: str, total_pending: float
) -> tuple[str, str]:
    tmpl = db.query(models.EmailTemplate).first()
    subject_t = tmpl.subject if tmpl else DEFAULT_EMAIL_SUBJECT
    body_t = tmpl.body if tmpl else DEFAULT_EMAIL_BODY

    # Ensure required placeholders exist so the recipient always
    # sees the rendidor name and the total pending amount, even if
    # the template stored in DB was edited and lost them.
    if "{{rendidor_nombre}}" not in subject_t and "{{rendidor_nombre}}" not in body_t:
        body_t = (
            f"Rendidor: {{rendidor_nombre}}\n\n{body_t}"
            if body_t
            else "Rendidor: {{rendidor_nombre}}"
        )
    if "{{monto_total_pendiente}}" not in subject_t and "{{monto_total_pendiente}}" not in body_t:
        extra_line = (
            "El monto total de las rendiciones pendientes "
            "de este informe es {{monto_total_pendiente}}."
        )
        body_t = f"{body_t}\n\n{extra_line}" if body_t else extra_line

    formatted_total = format_chilean_peso(total_pending)

    replacements = {
        "{{rendidor_nombre}}": rendidor_name,
        "{{monto_total_pendiente}}": formatted_total,
    }
    for key, value in replacements.items():
        subject_t = subject_t.replace(key, value)
        body_t = body_t.replace(key, value)
    return subject_t, body_t


class ConnectionManager:
    def __init__(self):
        self.active: set[WebSocket] = set()

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active.add(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active.discard(websocket)

    async def broadcast(self, message: dict):
        payload = json.dumps(message)
        for connection in list(self.active):
            try:
                await connection.send_text(payload)
            except Exception:
                self.disconnect(connection)


manager = ConnectionManager()


@app.post("/expenses", response_model=schemas.Expense)
async def create_expense(
    user_id: int = Form(...),
    title: str = Form(...),
    provider: str = Form(...),
    document_number: str = Form(...),
    document_type: str = Form(...),
    rendition_type: str = Form(...),
    description: str = Form(...),
    expense_date: str = Form(...),
    amount: float = Form(...),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
):
    normalized_type = document_type.upper()
    if normalized_type not in DOCUMENT_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de documento inválido")

    normalized_rendition = rendition_type.upper()
    if normalized_rendition not in RENDITION_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de rendición inválido")

    try:
        parsed_date = datetime.strptime(expense_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Fecha inválida. Usa formato YYYY-MM-DD")

    # Guardar archivo si viene
    document_path = None
    if file:
        ext = os.path.splitext(file.filename)[1]
        safe_title = "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in title)
        filename = f"{user_id}_{safe_title}_{uuid4().hex}{ext}"
        dest = UPLOAD_DIR / filename
        with dest.open("wb") as f:
            f.write(await file.read())
        document_path = f"/uploads/{filename}"

    expense_in = schemas.ExpenseCreate(
        user_id=user_id,
        title=title,
        provider=provider,
        document_number=document_number,
        document_type=normalized_type,
        rendition_type=normalized_rendition,
        description=description,
        expense_date=parsed_date,
        amount=amount,
    )
    try:
        expense = crud.create_expense(db, expense_in, document_path)
    except crud.BudgetError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not expense:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    await manager.broadcast({"type": "expense_created"})
    return expense


@app.get("/expenses", response_model=list[schemas.Expense])
def list_expenses(
    status: str | None = None,
    user_id: int | None = None,
    rendition_type: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    limit: int | None = None,
    offset: int | None = None,
    db: Session = Depends(get_db),
):
    parsed_start = parse_date_param(start_date, "start_date") if start_date else None
    parsed_end = parse_date_param(end_date, "end_date") if end_date else None
    if parsed_start and parsed_end and parsed_start > parsed_end:
        raise HTTPException(status_code=400, detail="start_date no puede ser mayor a end_date")
    if limit is not None:
        if limit <= 0 or limit > 100:
            raise HTTPException(status_code=400, detail="El límite debe estar entre 1 y 100")
    if offset is not None and offset < 0:
        raise HTTPException(status_code=400, detail="El offset debe ser positivo")
    return crud.get_expenses(
        db,
        status=status,
        user_id=user_id,
        rendition_type=rendition_type,
        start_date=parsed_start,
        end_date=parsed_end,
        limit=limit,
        offset=offset,
    )


@app.patch("/expenses/{expense_id}/status", response_model=schemas.Expense)
async def change_status(
    expense_id: int,
    status_update: schemas.ExpenseStatusUpdate,
    db: Session = Depends(get_db),
):
    try:
        exp = crud.update_expense_status(
            db, expense_id, status_update.status, status_update.approver_comment
        )
    except crud.BudgetError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not exp:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    await manager.broadcast({"type": "expense_updated"})
    return exp


@app.delete("/expenses/{expense_id}", status_code=204)
async def delete_expense(expense_id: int, db: Session = Depends(get_db)):
    deleted = crud.delete_expense(db, expense_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    await manager.broadcast({"type": "expense_deleted", "id": expense_id})


@app.get("/reports/summary", response_model=schemas.ReportSummary)
def report_summary(
    branch: str | None = None,
    user_id: int | None = None,
    db: Session = Depends(get_db),
):
    if branch:
        validate_branch(branch)
    return crud.get_report_summary(db, branch=branch, user_id=user_id)


@app.post("/reports/send-email")
def send_report_email(payload: ReportEmailRequest, db: Session = Depends(get_db)):
    if payload.branch:
        validate_branch(payload.branch)

    parsed_start = parse_date_param(payload.start_date, "start_date") if payload.start_date else None
    parsed_end = parse_date_param(payload.end_date, "end_date") if payload.end_date else None
    if parsed_start and parsed_end and parsed_start > parsed_end:
        raise HTTPException(status_code=400, detail="start_date no puede ser mayor a end_date")

    rendidor = crud.get_user(db, payload.user_id)
    if not rendidor:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    expenses = crud.get_expenses(
        db,
        status="PENDIENTE",
        user_id=payload.user_id,
        start_date=parsed_start,
        end_date=parsed_end,
    )
    if payload.branch:
        expenses = [e for e in expenses if e.branch == payload.branch]

    if payload.rendition_type == "FONDO_POR_RENDIR":
        # Solo fondo por rendir
        expenses = [e for e in expenses if e.rendition_type == "FONDO_POR_RENDIR"]
        is_fondo = True
    else:
        # Informe normal: excluimos fondo por rendir
        expenses = [e for e in expenses if e.rendition_type != "FONDO_POR_RENDIR"]
        is_fondo = False
    if not expenses:
        raise HTTPException(
            status_code=400,
            detail="No hay rendiciones pendientes para los filtros seleccionados.",
        )

    excel_content = _build_pending_excel(expenses, rendidor, is_fondo_por_rendir=is_fondo)
    total_pending = sum(float(e.amount or 0) for e in expenses)
    today_str = datetime.now().date().isoformat()
    safe_name = "".join(
        c if c.isalnum() or c in (" ", "_", "-") else "_"
        for c in rendidor.name
    ).strip()
    filename = f"{safe_name}_{today_str}"
    rendidor_name = rendidor.name
    subject, body = _render_email_template(db, rendidor_name, total_pending)
    _send_report_email(excel_content, filename, subject, body)
    return {"detail": "Correo enviado correctamente"}


@app.get("/reports/pending-excel")
def get_pending_excel(
    user_id: int,
    branch: str | None = None,
    rendition_type: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
):
    """
    Informe Excel único:
    - Si `rendition_type == FONDO_POR_RENDIR`: solo pendientes de fondo por rendir.
    - En cualquier otro caso: pendientes normales (excluye fondo por rendir).
    """
    if branch:
        validate_branch(branch)
    if rendition_type and rendition_type not in RENDITION_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de rendición inválido")

    parsed_start = parse_date_param(start_date, "start_date") if start_date else None
    parsed_end = parse_date_param(end_date, "end_date") if end_date else None
    if parsed_start and parsed_end and parsed_start > parsed_end:
        raise HTTPException(status_code=400, detail="start_date no puede ser mayor a end_date")

    user = crud.get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Siempre pendientes
    expenses = crud.get_expenses(
        db,
        status="PENDIENTE",
        user_id=user_id,
        start_date=parsed_start,
        end_date=parsed_end,
    )
    if branch:
        expenses = [e for e in expenses if e.branch == branch]

    if rendition_type == "FONDO_POR_RENDIR":
        # Solo fondo por rendir
        expenses = [e for e in expenses if e.rendition_type == "FONDO_POR_RENDIR"]
        is_fondo = True
    else:
        # Informe normal: excluimos fondo por rendir
        expenses = [e for e in expenses if e.rendition_type != "FONDO_POR_RENDIR"]
        is_fondo = False

    if not expenses:
        raise HTTPException(
            status_code=400,
            detail="No hay rendiciones pendientes para los filtros seleccionados.",
        )

    content = _build_pending_excel(expenses, user, is_fondo_por_rendir=is_fondo)
    today_str = datetime.now().date().isoformat()
    safe_name = "".join(
        c if c.isalnum() or c in (" ", "_", "-") else "_"
        for c in user.name
    ).strip()
    filename = f"{safe_name}_{today_str}"

    headers = {
        "Content-Disposition": f'attachment; filename=\"{filename}.xlsx\"'
    }
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/reports/pending-excel-debug")
def debug_pending_excel():
    info = {
        "template_path_normal": str(EXCEL_TEMPLATE_PATH),
        "template_exists_normal": EXCEL_TEMPLATE_PATH.exists(),
    }
    if info["template_exists_normal"]:
        try:
            wb = load_workbook(EXCEL_TEMPLATE_PATH)
            info["normal_sheetnames"] = wb.sheetnames
        except Exception as exc:
            info["normal_load_error"] = str(exc)
    return info


@app.get("/settings/email-template", response_model=schemas.EmailTemplate)
def get_email_template(db: Session = Depends(get_db)):
    tmpl = db.query(models.EmailTemplate).first()
    if not tmpl:
        return schemas.EmailTemplate(
            subject=DEFAULT_EMAIL_SUBJECT,
            body=DEFAULT_EMAIL_BODY,
        )
    return schemas.EmailTemplate(subject=tmpl.subject, body=tmpl.body)


@app.put("/settings/email-template", response_model=schemas.EmailTemplate)
def update_email_template(template_in: schemas.EmailTemplate, db: Session = Depends(get_db)):
    tmpl = db.query(models.EmailTemplate).first()
    if not tmpl:
        tmpl = models.EmailTemplate(subject=template_in.subject, body=template_in.body)
        db.add(tmpl)
    else:
        tmpl.subject = template_in.subject
        tmpl.body = template_in.body
    db.commit()
    return schemas.EmailTemplate(subject=tmpl.subject, body=tmpl.body)


@app.get("/users", response_model=list[schemas.User])
def list_users(
    role: str | None = None,
    branch: str | None = None,
    db: Session = Depends(get_db),
):
    if role:
        validate_role(role)
    if branch:
        validate_branch(branch)
    return crud.list_users(db, role=role, branch=branch)


@app.post("/users", response_model=schemas.User, status_code=201)
def create_user(user_in: schemas.UserCreate, db: Session = Depends(get_db)):
    validate_role(user_in.role)
    validate_branch(user_in.branch)
    existing = crud.get_user_by_email(db, user_in.email)
    if existing:
        raise HTTPException(status_code=409, detail="El correo ya está registrado")
    return crud.create_user(db, user_in)


@app.get("/users/{user_id}", response_model=schemas.User)
def get_user(user_id: int, db: Session = Depends(get_db)):
    user = crud.get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return user


@app.patch("/users/{user_id}", response_model=schemas.User)
def update_user(user_id: int, user_update: schemas.UserUpdate, db: Session = Depends(get_db)):
    if user_update.role:
        validate_role(user_update.role)
    if user_update.email:
        existing = crud.get_user_by_email(db, user_update.email)
        if existing and existing.id != user_id:
            raise HTTPException(status_code=409, detail="El correo ya está registrado por otro usuario")
    if user_update.branch:
        validate_branch(user_update.branch)
    user = crud.update_user(db, user_id, user_update)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return user


@app.delete("/users/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db)):
    deleted = crud.delete_user(db, user_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")


@app.post("/auth/login", response_model=schemas.User)
def login(credentials: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = crud.authenticate_user(db, credentials.email, credentials.password)
    if not user:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    return user


@app.post("/users/{user_id}/settlements", response_model=schemas.Movement, status_code=201)
async def create_settlement(
    user_id: int,
    payload: schemas.SettlementCreate,
    db: Session = Depends(get_db),
):
    try:
        movement = crud.settle_user_balance(db, user_id, payload.description)
    except crud.BudgetError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not movement:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    await manager.broadcast({"type": "settlement_created", "user_id": user_id})
    return movement


@app.get("/users/{user_id}/settlements", response_model=list[schemas.Movement])
def list_settlements(user_id: int, db: Session = Depends(get_db)):
    user = crud.get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return crud.list_movements(db, user_id)


@app.delete("/settlements/{movement_id}", status_code=204)
def delete_settlement(movement_id: int, db: Session = Depends(get_db)):
    deleted = crud.delete_movement(db, movement_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")


@app.websocket("/ws/approvals")
async def approvals_socket(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)
